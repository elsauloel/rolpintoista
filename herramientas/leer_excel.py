# -*- coding: utf-8 -*-
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from rutas import FICHA, GM, VENDOR, XLSX, SP

# Lee el Excel y reconstruye los items del catalogo. Se usa para comparar
# contra el codigo antes de aplicar nada.
from openpyxl import load_workbook


ETIQ = {
    'ID (no tocar)':'id','Nombre':'nombre','Tier':'tier','Tipo de ítem':'tipoItem',
    'Peso':'peso','Ranuras':'ranuras','Precio (DDE)':'precioCompra','Detalle':'detalle',
    'Tipo de dado':'tipoDado','Daño fijo':'danoFijo','Unidades':'unidades','Cargas':'cargaMax',
    'Cura HP':'curahp','Cura Bonos (%)':'curabonosPct','Estado que aplica':'efectoNombre',
    'Turnos del estado':'efectoTurnos','HP por turno':'efectoHpTurno',
    'Estado permanente':'efectoPermanente','Detalle del estado':'efectoDetalle',
    'Defensa':'mod_def','Res.Crít Tipo 2':'mod_tipo1','Res.Crít Tipo 4':'mod_tipo2',
    'Res.Crít Tipo 6':'mod_tipo3','Res.Crít Tipo 8':'mod_tipo4','Res.Crít Tipo 10':'mod_tipo5',
    'Movimiento':'mod_mov','Bonos':'mod_bonos','Evasión':'mod_eva','Iniciativa':'mod_ini',
    'PdG':'mod_pdg','Crítico %':'mod_crit','Parry':'mod_parry','Fuerza':'mod_fue',
    'Constitución':'mod_con','Res. mágica':'mod_resm','Res. CC':'mod_rescc',
    'Rango casteo':'mod_rangocasteo','Acciones máx.':'mod_accionesmax','Bloqueo':'mod_bloqueo',
    'Inteligencia':'mod_int','Destreza':'mod_des','Agilidad':'mod_agl',
    'Otros modificadores':'mods_extra','Tiene imagen':'tiene_imagen',
}

def num(v):
    if v in (None, ''): return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return float(v)
        except (TypeError, ValueError): return 0

def leer():
    wb = load_workbook(XLSX, data_only=True)
    items = []
    avisos = []
    for hoja in ['Consumibles','Armas','Escudos','Defensivos','Otros']:
        if hoja not in wb.sheetnames:
            avisos.append(f"Falta la pestaña {hoja}")
            continue
        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        if not filas: continue
        desconocidas = [c for c in filas[0] if c and c not in ETIQ]
        for c in desconocidas:
            avisos.append(f"{hoja}: columna desconocida {c!r} (se ignora)")
        cab = [ETIQ.get(c, None) for c in filas[0]]
        for nfila, f in enumerate(filas[1:], start=2):
            d = {}
            for k, v in zip(cab, f):
                if k: d[k] = v
            if not d.get('nombre'): continue          # fila vacía
            it = {'_hoja': hoja, '_fila': nfila, 'mods': []}
            for k, v in d.items():
                if k == 'tiene_imagen': continue
                if k == 'mods_extra':
                    if v:
                        for par in str(v).split(';'):
                            if ':' in par:
                                s, val = par.split(':', 1)
                                s = s.strip()
                                if s: it['mods'].append({'stat': s, 'val': num(val)})
                            elif par.strip():
                                avisos.append(f"{hoja} fila {nfila}: no entiendo {par.strip()!r} en Otros modificadores")
                    continue
                if k.startswith('mod_'):
                    if v not in (None, '') and num(v) != 0:
                        it['mods'].append({'stat': k[4:], 'val': num(v)})
                    continue
                if k == 'efectoPermanente':
                    it[k] = str(v).strip().lower() in ('sí','si','true','x','1') if v else False
                    continue
                if k in ('peso','ranuras','precioCompra','tipoDado','danoFijo','unidades',
                         'cargaMax','curahp','curabonosPct','efectoTurnos','efectoHpTurno'):
                    it[k] = num(v)
                    continue
                it[k] = (v if v is not None else '')
            it['consumible'] = (hoja == 'Consumibles')
            items.append(it)
    return items, avisos

if __name__ == '__main__':
    items, avisos = leer()
    print("Ítems leídos:", len(items))
    for a in avisos: print("  AVISO:", a)
