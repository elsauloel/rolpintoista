# -*- coding: utf-8 -*-
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from rutas import FICHA, GM, VENDOR, XLSX, SP

# Vuelca el catalogo a un Excel con una pestaña por categoria, listo para
# editar a mano y reimportar. Las imagenes no van (son base64 de cientos
# de KB): se conservan del catalogo actual al reimportar, por id.
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from catalogo_comun import CATALOGO_JSON


items = json.load(open(CATALOGO_JSON, encoding="utf-8"))
for it in items:
    it['_tiene_imagen'] = bool(it.get('imagen'))

DEFENSIVOS = {'armadura_blanda','armadura_rigida','casco_blando','casco_rigido',
              'guantes_blandos','guantes_rigidos','piernas_blandas','piernas_rigidas',
              'botas_blandas','botas_rigidas','cinturon'}
ACCESORIOS = {'anillos'}

def categoria(it):
    t = it['tipoItem']
    if it.get('consumible') or t == 'consumibles': return 'Consumibles'
    if t.startswith('arma_'): return 'Armas'
    if t.startswith('escudo'): return 'Escudos'
    if t in ACCESORIOS: return 'Accesorios'
    if t in DEFENSIVOS: return 'Defensivos'
    return 'Otros'

# --- Columnas por pestaña ---
BASE = ['id','nombre','tier','tipoItem','peso','ranuras','precioCompra','detalle']
CRIT = ['tipo1','tipo2','tipo3','tipo4','tipo5']
OTROS_MODS = ['mov','bonos','eva','ini','pdg','crit','parry','fue','con','int','agl','des',
              'resm','rescc','rangocasteo','accionesmax','dmg','potencia','hpmax','crgmax','bloqueo','rng','pdgmg']

COLUMNAS = {
    'Consumibles': BASE + ['legacy','unidades','cargaMax','curahp','curabonosPct',
                           'efectoNombre','efectoTurnos','efectoHpTurno','efectoPermanente','efectoDetalle','efectoMods_extra'],
    'Armas':       BASE + ['tipoDado','danoFijo','danoAmplificado','armaDeRango'] + ['mod_'+s for s in ['pdg','crit','ini','parry','fue','bloqueo']] + ['equipoEstadoNombre','equipoEstadoHpTurno','equipoEstadoDetalle'],
    'Escudos':     BASE + ['tipoDado','mod_def'] + ['mod_'+s for s in CRIT] + ['mod_parry'] + ['equipoEstadoNombre','equipoEstadoHpTurno','equipoEstadoDetalle'],
    'Defensivos':  BASE + ['mod_def'] + ['mod_'+s for s in CRIT] +
                   ['mod_'+s for s in ['mov','bonos','eva','ini','pdg','parry','fue','con','int','des','agl','resm','rescc','rangocasteo','accionesmax']] + ['equipoEstadoNombre','equipoEstadoHpTurno','equipoEstadoDetalle'],
    'Accesorios':  BASE + ['tipoDado','danoFijo','unidades','cargaMax','curahp','curabonosPct','mod_def',
                           'equipoEstadoNombre','equipoEstadoHpTurno','equipoEstadoDetalle','equipoEstadoPreset'],
    'Otros':       BASE + ['tipoDado','danoFijo','unidades','cargaMax','curahp','curabonosPct','mod_def'],
}
# Al final de cada pestaña: lo que no tiene columna propia y el aviso de imagen
for k in COLUMNAS:
    COLUMNAS[k] = COLUMNAS[k] + ['mods_extra','tiene_imagen']

ETIQUETAS = {
    'id':'ID (no tocar)','nombre':'Nombre','tier':'Tier','tipoItem':'Tipo de ítem',
    'peso':'Peso','ranuras':'Ranuras','precioCompra':'Precio (DDE)','detalle':'Detalle',
    'tipoDado':'Tipo de dado','danoFijo':'Daño fijo','danoAmplificado':'Daño amplificado','armaDeRango':'Arma de rango',
    'unidades':'Unidades','cargaMax':'Cargas',
    'curahp':'Cura HP','curabonosPct':'Cura Bonos (%)',
    'efectoNombre':'Estado que aplica','efectoTurnos':'Turnos del estado',
    'efectoHpTurno':'HP por turno','efectoPermanente':'Estado permanente','efectoDetalle':'Detalle del estado',
    'efectoMods_extra':'Modificadores del estado',
    'equipoEstadoNombre':'Estado al equipar','equipoEstadoHpTurno':'HP por turno equipado',
    'equipoEstadoDetalle':'Detalle del estado equipado',
    'equipoEstadoPreset':'Estado al equipar (preset)',
    'legacy':'Legacy',
    'mods_extra':'Otros modificadores','tiene_imagen':'Tiene imagen',
    'mod_def':'Defensa','mod_tipo1':'Res.Crít Tipo 2','mod_tipo2':'Res.Crít Tipo 4',
    'mod_tipo3':'Res.Crít Tipo 6','mod_tipo4':'Res.Crít Tipo 8','mod_tipo5':'Res.Crít Tipo 10',
    'mod_mov':'Movimiento','mod_bonos':'Bonos','mod_eva':'Evasión','mod_ini':'Iniciativa',
    'mod_pdg':'PdG','mod_crit':'Crítico %','mod_parry':'Parry','mod_fue':'Fuerza','mod_con':'Constitución',
    'mod_resm':'Res. mental','mod_rescc':'Res. CC','mod_rangocasteo':'Rango casteo',
    'mod_accionesmax':'Acciones máx.','mod_bloqueo':'Bloqueo','mod_int':'Inteligencia','mod_des':'Destreza','mod_agl':'Agilidad',
}

def valor(it, col):
    if col == 'tiene_imagen':
        return 'sí' if it.get('_tiene_imagen') else ''
    if col == 'mods_extra':
        propios = {c[4:] for c in COLUMNAS[categoria(it)] if c.startswith('mod_')}
        extra = [f"{m['stat']}:{m['val']}" for m in it.get('mods', []) if m['stat'] not in propios]
        return '; '.join(extra)
    if col == 'efectoMods_extra':
        return '; '.join(f"{m['stat']}:{m['val']}" for m in it.get('efectoMods', []))
    if col.startswith('mod_'):
        stat = col[4:]
        for m in it.get('mods', []):
            if m['stat'] == stat:
                return m['val']
        return ''
    if col == 'legacy':
        return 'si' if it.get('legacy') else ''
    v = it.get(col, '')
    if isinstance(v, bool):
        return 'sí' if v else ''
    return v

wb = Workbook()
wb.remove(wb.active)

CAB = PatternFill('solid', fgColor='2E1F16')
CAB_ID = PatternFill('solid', fgColor='4A2E1C')
FUENTE = Font(bold=True, color='F0E4D0', size=10)

resumen = {}
for cat in ['Consumibles','Armas','Escudos','Defensivos','Accesorios','Otros']:
    ws = wb.create_sheet(cat)
    cols = COLUMNAS[cat]
    ws.append([ETIQUETAS.get(c, c) for c in cols])
    for i, c in enumerate(cols, 1):
        celda = ws.cell(row=1, column=i)
        celda.fill = CAB_ID if c in ('id','tiene_imagen') else CAB
        celda.font = FUENTE
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ancho = 34 if c == 'detalle' else 30 if c in ('efectoDetalle','mods_extra') else \
                26 if c == 'nombre' else 22 if c in ('id','tipoItem','efectoNombre') else 13
        ws.column_dimensions[get_column_letter(i)].width = ancho

    delCat = sorted([it for it in items if categoria(it) == cat],
                    key=lambda x: (x['tier'], x['nombre']))
    for it in delCat:
        ws.append([valor(it, c) for c in cols])
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(1, len(delCat)) + 1}"
    resumen[cat] = len(delCat)

# --- Hoja de referencia ---
ws = wb.create_sheet('Referencia')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 96
filas = [
    ('CÓMO USAR ESTE ARCHIVO', ''),
    ('', 'Este Excel es la fuente editable del catálogo. Se edita acá y después se reimporta al programa.'),
    ('', ''),
    ('ID (no tocar)', 'Identifica al ítem. Es lo que permite reconocerlo al reimportar. Si lo cambiás, el programa lo toma como un ítem nuevo.'),
    ('Ítems nuevos', 'Agregalos en la fila que sigue, en la pestaña que corresponda. Dejá el ID vacío: se genera solo al importar.'),
    ('Borrar un ítem', 'Borrá la fila entera. Al reimportar deja de estar en el catálogo.'),
    ('', ''),
    ('Tier', 'Común · Buena Calidad · Raro · Excepcional · Legendario. Se escribe igual, con las mayúsculas puestas.'),
    ('Tipo de ítem', 'Define cómo se comporta y en qué ranura entra. Valores válidos: arma_1m, arma_2m, escudo_1m, escudo_2m, armadura_blanda, armadura_rigida, casco_blando, casco_rigido, guantes_blandos, guantes_rigidos, piernas_blandas, piernas_rigidas, botas_blandas, botas_rigidas, consumibles, anillos, otros.'),
    ('Peso', 'En las armas es además la cantidad de dados que tira: peso 3 con dado d6 = 3d6.'),
    ('Tipo de dado', 'El número del dado del arma: 2, 4, 6, 8 o 10.'),
    ('Daño fijo', 'Se suma al resultado de los dados.'),
    ('', ''),
    ('Columnas de modificadores', 'Cada una suma (o resta, con negativo) a ese stat mientras el ítem esté equipado. Vacío = no lo toca.'),
    ('Otros modificadores', 'Para un stat que no tiene columna propia en esa pestaña. Formato: stat:valor; stat:valor — por ejemplo  mov:1; bonos:2'),
    ('', ''),
    ('Consumibles', 'Unidades es cuántas trae. Cargas es cuántos usos por unidad. Cura HP suma vida al consumirlo (negativo daña).'),
    ('Estado que aplica', 'Si tiene nombre, al consumir el ítem se activa ese estado alterado. Turnos, HP por turno y Detalle configuran ese estado.'),
    ('Estado permanente', 'Con "sí", el estado no vence: queda hasta que el jugador lo borre a mano.'),
    ('', ''),
    ('Estado al equipar', 'Igual que "Estado que aplica" pero para equipo (armas, escudos, defensivos, accesorios): el estado se activa solo mientras el ítem está puesto y se va al sacárselo — no cuenta turnos.'),
    ('Estado al equipar (preset)', 'Opcional. Si el nombre de un preset de la ficha (Afortunado, Sangre pura, etc.) va acá, el ítem hereda esa mecánica real además de mostrar su propio nombre en "Estado al equipar" — así "Afortunado (anillo)" da ventaja de verdad en vez de ser solo un cartel. Dejalo vacío si el efecto no tiene un preset que lo cubra: el estado va a aparecer igual, como recordatorio, pero sin aplicarse solo.'),
    ('', ''),
    ('Tiene imagen', 'Solo informativo. Las imágenes NO están en este Excel porque pesan cientos de KB cada una; viven en el programa y se conservan solas al reimportar, siempre que no cambies el ID.'),
    ('', ''),
    ('CONTENIDO ACTUAL', ''),
]
for c, n in resumen.items():
    filas.append((c, f'{n} ítems'))
filas.append(('TOTAL', f'{sum(resumen.values())} ítems'))

for a, b in filas:
    ws.append([a, b])
for fila in ws.iter_rows():
    for c in fila:
        c.alignment = Alignment(vertical='top', wrap_text=True)
        if c.column == 1 and c.value and str(c.value).isupper():
            c.font = Font(bold=True, size=11)
        elif c.column == 1 and c.value:
            c.font = Font(bold=True, size=10)

wb.save(XLSX)
print('Generado:', XLSX)
for c, n in resumen.items():
    print(f'  {c}: {n}')
print('  TOTAL:', sum(resumen.values()))
